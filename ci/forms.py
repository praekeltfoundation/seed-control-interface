from django import forms
from django.conf import settings
from django.utils.translation import ugettext_lazy as _
from seed_services_client.auth import AuthApiClient
from demands import HTTPServiceError
from bootstrap_datepicker.widgets import DatePicker
from django.contrib.postgres.forms import SimpleArrayField
from openpyxl import load_workbook


class AuthenticationForm(forms.Form):
    """
    Base class for authenticating users. Extend this to get a form that accepts
    username/password logins.
    """
    username = forms.CharField(max_length=254, label=_("Email"))
    password = forms.CharField(label=_("Password"), widget=forms.PasswordInput)

    error_messages = {
        'invalid_login': _("Please enter a correct %(username)s and password. "
                           "Note that both fields may be case-sensitive."),
        'inactive': _("This account is inactive."),
    }

    def __init__(self, request=None, *args, **kwargs):
        """
        The 'request' parameter is set for custom auth use by subclasses.
        The form data comes in via the standard 'data' kwarg.
        """
        self.request = request
        self.user_cache = None
        super(AuthenticationForm, self).__init__(*args, **kwargs)

    def clean(self):
        username = self.cleaned_data.get('username')
        password = self.cleaned_data.get('password')

        if username and password:
            # do remote login
            auth_url = settings.AUTH_SERVICE_URL
            try:
                auth_api = AuthApiClient(username, password, auth_url)
            except HTTPServiceError:
                raise forms.ValidationError(
                    self.error_messages['invalid_login'],
                    code='invalid_login',
                    params={'username': 'Email'},
                )
            self.user_cache = auth_api.get_permissions()
            self.user_cache["token"] = auth_api.token

        return self.cleaned_data

    def get_user(self):
        return self.user_cache


ADDRESS_TYPES = (
    ("msisdn", "Cell Phone"),
)


class IdentitySearchForm(forms.Form):
    address_value = forms.CharField(
        widget=forms.TextInput(attrs={'placeholder': 'Search for...'}))
    address_type = forms.ChoiceField(
        choices=ADDRESS_TYPES
    )


class AddSubscriptionForm(forms.Form):
    messageset = forms.IntegerField()


class DeactivateSubscriptionForm(forms.Form):
    subscription_id = forms.CharField()


class ChangeSubscriptionForm(forms.Form):
    language = forms.CharField()
    messageset = forms.IntegerField()

default_stage = (
    ('', "Stage - all"),
)
STAGE_CHOICES = default_stage + settings.STAGES

VALIDATED_CHOICES = (
    ('', "Validated - all"),
    ("True", "Valid"),
    ("False", "Invalid")
)


class RegistrationFilterForm(forms.Form):
    mother_id = forms.CharField(
        widget=forms.TextInput(attrs={'placeholder': 'Mother ID'}),
        required=False)
    stage = forms.ChoiceField(
        choices=STAGE_CHOICES, required=False
    )
    validated = forms.ChoiceField(
        choices=VALIDATED_CHOICES, required=False
    )


default_action = (
    ('', "Action - all"),
)
ACTION_CHOICES = default_action + settings.ACTIONS


class ChangeFilterForm(forms.Form):
    mother_id = forms.CharField(
        widget=forms.TextInput(attrs={'placeholder': 'Mother ID'}),
        required=False)
    action = forms.ChoiceField(
        choices=ACTION_CHOICES, required=False
    )
    validated = forms.ChoiceField(
        choices=VALIDATED_CHOICES, required=False
    )


ACTIVE_CHOICES = (
    ('', "Active - all"),
    ("True", "Active"),
    ("False", "Inactive")
)

COMPLETED_CHOICES = (
    ('', "Completed - all"),
    ("True", "Completed"),
    ("False", "Incompleted")
)

REPORT_CHOICES = (
    ("registration", "Registration Report"),
    ("cohort", "Cohort Report")
)


class SubscriptionFilterForm(forms.Form):
    identity = forms.CharField(
        widget=forms.TextInput(attrs={'placeholder': 'Mother ID'}),
        required=False)
    active = forms.ChoiceField(
        choices=ACTIVE_CHOICES, required=False
    )
    completed = forms.ChoiceField(
        choices=COMPLETED_CHOICES, required=False
    )


class ReportGenerationForm(forms.Form):
    report_type = forms.ChoiceField(
        choices=REPORT_CHOICES, widget=forms.HiddenInput(), required=True)
    start_date = forms.DateField(
        widget=DatePicker(options={
            'placeholder': 'YYYY-MM-DD', 'format': 'yyyy-mm-dd',
            'autoclose': True}),
        required=False)
    end_date = forms.DateField(
        widget=DatePicker(options={
            'placeholder': 'YYYY-MM-DD', 'format': 'yyyy-mm-dd',
            'autoclose': True}),
        required=False)
    email_to = SimpleArrayField(forms.EmailField(), required=True)
    email_from = forms.EmailField(
        widget=forms.EmailInput(attrs={
            'placeholder': 'Address the email is from'}),
        required=False)
    email_subject = forms.CharField(
        widget=forms.TextInput(attrs={'placeholder': 'Subject for email'}),
        required=False)


class MsisdnReportGenerationForm(ReportGenerationForm):
    msisdn_list = forms.FileField()

    def clean_msisdn_list(self):
        msisdns = []
        wb = load_workbook(self.cleaned_data['msisdn_list'])
        ws = wb[wb.sheetnames[0]]  # There should only be one sheet
        # Get msisdn column
        msisdn_column = None
        for i, column in enumerate(ws.columns):
            if column[0].value is None:
                # Skip untitled columns
                continue
            value = str(column[0].value)
            if value.lower() == 'phone number':
                msisdn_column = i
                break
        if msisdn_column is None:
            raise forms.ValidationError(
                "Invalid contents for: %(file)s. File must contain "
                "'Phone number' column in the first sheet",
                code='invalid',
                params={'file': self.cleaned_data['msisdn_list']})

        for row in ws.rows:
            value = row[msisdn_column].value
            if value is None:
                # Skip empty Rows
                continue

            value = str(value)
            if value.lower() == 'phone number':
                # Skip the header row
                continue

            # Clean the msisdn
            msisdn = value.replace(" ", "")
            if not msisdn.replace("+", "").isdigit():
                raise forms.ValidationError(
                    "Invalid contents for: %(file)s. 'Phone number' column "
                    "must only contain valid phone numbers",
                    code='invalid',
                    params={'file': self.cleaned_data['msisdn_list']})
            msisdns.append(msisdn)
        return msisdns
